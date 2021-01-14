#!/usr/bin/env python
# -*- coding: utf-8 -*-
# encoding: utf-8

'''
# @Time    : 2020/11/20  8:47 下午
# @Author  : yantianpeng
# @Site    : 
# @File    : read_openpyxl.py
# @Software: PyCharm
'''
from openpyxl import load_workbook
class Openxls:

    def __init__(self,file,sheetname):
        self.file = file;
        self.sheetname = sheetname;

    def readContent(self):
        #加载文件
        wb = load_workbook(filename=self.file);
        #加载sheet
        if self.sheetname is None:
            ws =  wb.active
        else:
            ws =  wb[self.sheetname];
        #创建一个空的列表
        row = ws.max_row;#最大行
        column = ws.max_column#最大列
        data = [];#用来存放组装的数组
        title ={};#存放每一行的数据，作为一个字典,用于后面的转换
        #先读取第一行的数据，作为字典的健
        for i in range(1,column+1):
            title[i] = ws.cell(1,i).value;#字典的新增

        for j in range(2,row+1):
            temp ={};#临时变量用来存放字典
            for i in range(1,column+1):
                temp[title[i]] = str(ws.cell(row=j,column=i).value);#字典的新增,这一步需要把获取的值转换为字符串,因为整数的话是没有办法比较长度的
            temp[title[i]] = eval(temp[title[i]]);#把expect字符串转换为字典
            data.append(temp);
        return data;



# if __name__ == '__main__':
#     data_01 = Openxls('case_01.xlsx','login').readContent();
#     print(type(data_01[0]['expect']));