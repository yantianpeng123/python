#!/usr/bin/env python
# -*- coding: utf-8 -*-
# encoding: utf-8

'''
# @Time    : 2020/12/1  7:04 下午
# @Author  : yantianpeng
# @Site    : 
# @File    : read_excle.py
# @Software: PyCharm
'''
from openpyxl import load_workbook;
import json

class OpenExcel:

    def __init__(self,file,sheetname):
        self.file = file;
        self.sheetname = sheetname;


    def open_read_excle(self):
        #加载文件
        wb = load_workbook(filename=self.file);
        #加载sheet
        if self.sheetname is None:
            ws = wb.active;
        else:
            ws = wb[self.sheetname];
        #获取最大行和最大列
        row =  ws.max_row;
        column = ws.max_column;
        data = [];#用来存放组装数据
        title = {};#存放每一行数据，用来作为字典的转换，作为一个临时变量
        for i in range(1,column+1):
            title[i] = ws.cell(1,i).value;

        for j in range(2,row+1):
            temp = {};#用来存储临时变量
            for i in range(1,column+1):
                temp[title[i]] = str(ws.cell(row=j,column=i).value);
            temp[title[i]]= eval(temp[title[i]]);
            data.append(temp);
        return data;

    def open_read_excle_str(self):
        #加载文件
        wb = load_workbook(filename=self.file);
        #加载sheet
        if self.sheetname is None:
            ws = wb.active;
        else:
            ws = wb[self.sheetname];
        #获取最大行和最大列
        row =  ws.max_row;
        column = ws.max_column;
        data = [];#用来存放组装数据
        title = {};#存放每一行数据，用来作为字典的转换，作为一个临时变量
        for i in range(1,column+1):
            title[i] = ws.cell(1,i).value;

        for j in range(2,row+1):
            temp = {};#用来存储临时变量
            for i in range(1,column+1):
                temp[title[i]] = (ws.cell(row=j,column=i).value);
            data.append(temp);
        return data;

if __name__ == '__main__':
    # oe = OpenExcel(file='../day_11/case_01.xlsx',sheetname='register');
    # data = oe.open_read_excle();
    # print(data);  ningmengban.xlsx
    oe = OpenExcel(file="",sheetname='register');
    data = oe.open_read_excle_str();

    print(data[0],type(data[0]))
    json_01 = json.dumps(data,ensure_ascii=False);#python转json
    print(json_01);
    # python_str = json.loads(json_01);
    # print(python_str);#json对象转python对象